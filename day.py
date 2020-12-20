from openpyxl import load_workbook

wb = load_workbook('./lxml/schedule.xlsx')
sheet = wb.active
var = sheet['A1'].value
c = sheet['B2']
c.row
c.column
c.coordinate
no_lesson = 'Сегодня пар нет! Можно почилить в doka2'

# ПОНЕДЕЛЬНИК
massive_monday = ['Понедельник']
lesson = 0
for monday in range(8):
    lesson += 1
    monday = sheet.cell(row=lesson, column=2).value
    time = sheet.cell(row=lesson, column=1).value
    if monday == 'Понедельник':
        continue
    if monday is None:  # избавились от пустых значений в таблице
        continue
    massive_monday.append(f'{lesson - 1} | {time} | ' + monday)
if len(massive_monday) == 1:
    massive_monday.append(no_lesson)
monday = '\n'.join(massive_monday)  # делать все через массив

# ВТОРНИК
massive_tuesday = ['Вторник']
lesson = 0
for tuesday in range(8):
    lesson += 1
    tuesday = sheet.cell(row=lesson, column=3).value
    time = sheet.cell(row=lesson, column=1).value
    if tuesday == 'Вторник':
        continue
    if tuesday is None:  # избавились от пустых значений в таблице
        continue
    massive_tuesday.append(f'{lesson - 1} | {time} | ' + tuesday)
if len(massive_tuesday) == 1:
    massive_tuesday.append(no_lesson)
tuesday = '\n'.join(massive_tuesday)  # делать все через массив

# СРЕДА
massive_wednesday = ['Среда']
lesson = 0
for wednesday in range(8):
    lesson += 1
    wednesday = sheet.cell(row=lesson, column=4).value
    time = sheet.cell(row=lesson, column=1).value
    if wednesday == 'Среда':
        continue
    if wednesday is None:  # избавились от пустых значений в таблице
        continue
    massive_wednesday.append(f'{lesson - 1} | {time} | ' + wednesday)
if len(massive_wednesday) == 1:
    massive_wednesday.append(no_lesson)
wednesday = '\n'.join(massive_wednesday)  # делать все через массив

# ЧЕТВЕРГ
massive_thursday = ['Четверг']
lesson = 0
for thursday in range(8):
    lesson += 1
    thursday = sheet.cell(row=lesson, column=5).value
    time = sheet.cell(row=lesson, column=1).value
    if thursday == 'Четверг':
        continue
    if thursday is None:  # избавились от пустых значений в таблице
        continue
    massive_thursday.append(f'{lesson - 1} | {time} | ' + thursday)
if len(massive_thursday) == 1:
    massive_thursday.append(no_lesson)
thursday = '\n'.join(massive_thursday)  # делать все через массив

# ПЯТНИЦА
massive_friday = ['Пятница']
lesson = 0
for friday in range(8):
    lesson += 1
    friday = sheet.cell(row=lesson, column=6).value
    time = sheet.cell(row=lesson, column=1).value
    if friday == 'Пятница':
        continue
    if friday is None:  # избавились от пустых значений в таблице
        continue
    massive_friday.append(f'{lesson - 1} | {time} | ' + friday)
if len(massive_friday) == 1:
    massive_friday.append(no_lesson)
friday = '\n'.join(massive_friday)  # делать все через массив

# СУББОТА
massive_saturday = ['Пятница']
lesson = 0
for saturday in range(8):
    lesson += 1
    saturday = sheet.cell(row=lesson, column=7).value
    time = sheet.cell(row=lesson, column=1).value
    if saturday == 'Пятница':
        continue
    if saturday is None:  # избавились от пустых значений в таблице
        continue
    massive_saturday.append(f'{lesson - 1} | {time} | ' + saturday)
if len(massive_saturday) == 1:
    massive_saturday.append(no_lesson)
saturday = '\n'.join(massive_saturday)  # делать все через массив

# ПАРЫ
massive_time = ['Расписание звонков:']
lesson = 0
for time in range(8):
    lesson += 1
    time = sheet.cell(row=lesson, column=1).value
    end_time = sheet.cell(row=lesson, column=8).value
    if time == 'Время':
        continue
    if time is None:  # избавились от пустых значений в таблице
        continue
    massive_time.append(f'{lesson - 1} {time} - {end_time}')
time = '\n'.join(massive_time)  # делать все через массив